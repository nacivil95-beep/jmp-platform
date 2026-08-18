# -*- coding: utf-8 -*-
"""
extract_board.py
─────────────────────────────────────────────────────────────────────────
데이터관리/알림게시판/board.xlsx (시트명: "게시판") 를 읽어 루트의 board.js를 생성한다.
대시보드 헤더의 "게시판" 티커(좌→우로 흐르는 텍스트)가 이 board.js의
BOARD_ITEMS 배열을 읽어 화면에 표시한다.

사용법 (레포 루트에서 실행):
        python3 extract_board.py
    → board.js 가 새로 생성된다. board.xlsx / board.js / extract_board.py를
      함께 GitHub에 push하면 사이트에 반영된다 (index.html이 board.js를 <script>로 로드).

엑셀 컬럼 규칙 (1행 헤더):
    id         : 고유 값 (문자열/숫자 무관)
    text       : 전광판에 흘러갈 문구 (줄바꿈 없는 한 줄)
    startDate  : "YYYY-MM-DD" (선택) - 이 날짜부터 노출
    endDate    : "YYYY-MM-DD" (선택) - 이 날짜까지 노출 (포함)
    pinned     : TRUE/FALSE (선택) - TRUE면 다른 항목보다 먼저 노출

A열이 '/*' 로 시작하는 행(엑셀 파일 내 설명 주석 블록)과, text가 비어있는
행은 자동으로 건너뛴다.
"""

import json
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl이 필요합니다. 설치: pip install openpyxl --break-system-packages")

ROOT = Path(__file__).resolve().parent
XLSX_PATH = ROOT / "데이터관리" / "알림게시판" / "board.xlsx"
JS_PATH = ROOT / "board.js"
SHEET_NAME = "게시판"


def normalize_date(value):
    """엑셀 셀 값을 'YYYY-MM-DD' 문자열 또는 None으로 정규화."""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    return text or None


def normalize_bool(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in ("true", "1", "y", "yes", "참")


def is_comment_row(id_value):
    """엑셀에 남겨둔 '/** ... */' 설명 블록 행인지 판별."""
    if id_value is None:
        return False
    text = str(id_value).strip()
    return text.startswith("/*") or text.startswith("*") or text.startswith("*/")


def extract_items():
    if not XLSX_PATH.exists():
        sys.exit(f"파일을 찾을 수 없습니다: {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb.worksheets[0]  # 시트명이 다르면 첫 번째 시트를 사용

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return []

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_index = {name: idx for idx, name in enumerate(header)}

    required = ["id", "text"]
    missing = [c for c in required if c not in col_index]
    if missing:
        sys.exit(f"필수 컬럼이 없습니다: {missing} (헤더: {header})")

    items = []
    for row in rows[1:]:
        def get(col):
            idx = col_index.get(col)
            return row[idx] if idx is not None and idx < len(row) else None

        raw_id = get("id")
        text = get("text")

        if is_comment_row(raw_id):
            continue
        if text is None or not str(text).strip():
            continue

        items.append({
            "id": raw_id if raw_id is not None else len(items) + 1,
            "text": str(text).strip(),
            "startDate": normalize_date(get("startDate")),
            "endDate": normalize_date(get("endDate")),
            "pinned": normalize_bool(get("pinned")),
        })

    return items


def write_js(items):
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    body = json.dumps(items, ensure_ascii=False, indent=2)
    js = f"""// board.js — 자동 생성 파일입니다. 직접 수정하지 마세요.
// 원본: 데이터관리/알림게시판/board.xlsx (시트: "{SHEET_NAME}")
// 생성 스크립트: extract_board.py
// 생성 시각: {generated_at}
//
// 이 배열은 헤더의 "게시판" 티커(좌→우로 흐르는 텍스트)가 읽어 표시합니다.
// 내용을 바꾸려면 board.xlsx를 수정한 뒤 extract_board.py를 다시 실행하세요.

const BOARD_ITEMS = {body};
"""
    JS_PATH.write_text(js, encoding="utf-8")


def main():
    items = extract_items()
    write_js(items)
    print(f"[extract_board.py] {len(items)}개 항목 → {JS_PATH.name} 생성 완료")
    for it in items:
        print(f"  - #{it['id']} {it['text'][:40]}{'...' if len(it['text']) > 40 else ''}")


if __name__ == "__main__":
    main()
