import os
import re
import json
import datetime
import openpyxl


def to_num(val):
    if val is None:
        return 0
    try:
        n = float(val)
        if n.is_integer():
            return int(n)
        return n
    except (ValueError, TypeError):
        return 0


def to_str(val):
    if val is None:
        return ""
    return str(val).strip()


def parse_first_date(text):
    """'2026년 07월 13일 ~ 07월 14일' 같은 문자열에서 맨 앞 날짜만 뽑아 YYYY-MM-DD로 변환합니다."""
    if not text:
        return ""
    m = re.search(r"(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})", str(text))
    if not m:
        return ""
    y, mo, d = (int(x) for x in m.groups())
    try:
        return datetime.date(y, mo, d).strftime("%Y-%m-%d")
    except ValueError:
        return ""


def parse_range(text):
    """'11.5 ~17.5' -> (11.5, 17.5)"""
    if not text:
        return None, None
    m = re.match(r"\s*([\d.]+)\s*~\s*([\d.]+)", str(text))
    if not m:
        return None, None
    return float(m.group(1)), float(m.group(2))


def parse_tester_line(text):
    """'시험자 : 윤 정 필  (인)   확인자 : 백 귀 석  (인)   건설사업관리기술인 : 금 동 훈  (인)' 파싱"""
    result = {"tester": "", "checker": "", "manager": ""}
    if not text:
        return result
    patterns = {
        "tester": r"시험자\s*:\s*([^\(]+)",
        "checker": r"확인자\s*:\s*([^\(]+)",
        "manager": r"건설사업관리기술인\s*:\s*([^\(]+)"
    }
    for key, pat in patterns.items():
        m = re.search(pat, str(text))
        if m:
            result[key] = m.group(1).strip()
    return result


# =============================================================================
# 1) 다짐 전 함수비 시험 - 시트 하나당 시험 1건, 시험번호 = 시트명
# =============================================================================
def collect_moisture(wb):
    results = []
    for sheet_name in wb.sheetnames:
        if not sheet_name.isdigit():
            continue
        ws = wb[sheet_name]

        location = to_str(ws["M5"].value)
        date_raw = to_str(ws["M4"].value)
        judge = to_str(ws["F18"].value)
        if not location and not judge:
            continue

        avg = to_num(ws["F15"].value)
        range_min, range_max = parse_range(ws["F17"].value)

        samples = []
        for col in ["F", "I", "L"]:
            label = ws[f"{col}7"].value
            if label is None or str(label).strip() == "":
                continue
            samples.append({
                "label": to_str(label),
                "container": to_str(ws[f"{col}8"].value),
                "wet": to_num(ws[f"{col}9"].value),
                "dry": to_num(ws[f"{col}10"].value),
                "water": to_num(ws[f"{col}12"].value),
                "moisture": round(to_num(ws[f"{col}14"].value), 2)
            })

        tester = to_str(ws["C37"].value)
        checker = to_str(ws["I37"].value)
        manager = to_str(ws["M37"].value).replace("성   명 :", "").strip()

        results.append({
            "no": sheet_name,
            "date": parse_first_date(date_raw),
            "date_raw": date_raw,
            "location": location,
            "avg": round(avg, 2) if avg else 0,
            "range_min": range_min,
            "range_max": range_max,
            "judge": judge,
            "samples": samples,
            "testers": {"tester": tester, "checker": checker, "manager": manager}
        })

    results.sort(key=lambda r: int(r["no"]))
    return results


# =============================================================================
# 2) 평판재하시험 - "List" 시트에 전체 요약이 있고, 시트명과 같은 개별 시트에 상세 있음
# =============================================================================
def collect_plate(wb):
    if "List" not in wb.sheetnames:
        return []
    list_ws = wb["List"]

    results = []
    r = 4
    while True:
        no = list_ws[f"A{r}"].value
        if no is None:
            break
        date_val = list_ws[f"B{r}"].value
        location = to_str(list_ws[f"C{r}"].value)
        k30 = to_num(list_ws[f"D{r}"].value)
        standard = to_num(list_ws[f"E{r}"].value)
        judge = to_str(list_ws[f"F{r}"].value)

        date_str = ""
        if isinstance(date_val, (datetime.datetime, datetime.date)):
            date_str = date_val.strftime("%Y-%m-%d")

        detail_sheet_name = str(int(no)) if isinstance(no, (int, float)) else str(no)
        plate_diameter = plate_area = initial_load = conclusion = ""
        testers = {"tester": "", "checker": "", "manager": ""}
        if detail_sheet_name in wb.sheetnames:
            dws = wb[detail_sheet_name]
            plate_diameter = to_num(dws["C4"].value)
            plate_area = to_num(dws["E4"].value)
            initial_load = to_str(dws["G4"].value)
            conclusion = to_str(dws["H34"].value) or to_str(dws["H36"].value)
            testers = parse_tester_line(dws["A43"].value)

        results.append({
            "no": int(no) if isinstance(no, (int, float)) else no,
            "date": date_str,
            "location": location,
            "k30": k30,
            "standard": standard,
            "judge": judge,
            "plate_diameter": plate_diameter,
            "plate_area": plate_area,
            "initial_load": initial_load,
            "conclusion": conclusion,
            "testers": testers
        })
        r += 1

    results.sort(key=lambda x: x["no"] if isinstance(x["no"], (int, float)) else 0)
    return results


def extract_quality():
    quality_dir = "품질시험"
    if not os.path.exists(quality_dir):
        print(f"에러: '{quality_dir}' 폴더가 존재하지 않습니다.")
        return

    moisture_file = None
    plate_file = None
    for f in os.listdir(quality_dir):
        if not f.endswith(".xlsx") or f.startswith("~$"):
            continue
        if "함수비" in f:
            moisture_file = os.path.join(quality_dir, f)
        elif "평판재하" in f:
            plate_file = os.path.join(quality_dir, f)

    moisture_results, plate_results = [], []

    if moisture_file:
        print(f"함수비 시험 파일 읽는 중: {moisture_file}")
        wb = openpyxl.load_workbook(moisture_file, data_only=True)
        moisture_results = collect_moisture(wb)
        print(f"함수비 시험 {len(moisture_results)}건 파싱 완료")
    else:
        print("경고: '품질시험' 폴더에서 함수비 엑셀 파일을 찾을 수 없습니다.")

    if plate_file:
        print(f"평판재하시험 파일 읽는 중: {plate_file}")
        wb2 = openpyxl.load_workbook(plate_file, data_only=True)
        plate_results = collect_plate(wb2)
        print(f"평판재하시험 {len(plate_results)}건 파싱 완료")
    else:
        print("경고: '품질시험' 폴더에서 평판재하시험 엑셀 파일을 찾을 수 없습니다.")

    output_file = "qualityData.js"
    with open(output_file, "w", encoding="utf-8") as f:
        f.write("/**\n")
        f.write(" * 품질시험 엑셀 자동 추출 결과 데이터 (다짐전함수비 / 평판재하시험)\n")
        f.write(" * 생성 스크립트: extract_quality.py\n")
        f.write(" */\n")
        f.write("let QUALITY_DATA = ")
        json.dump({"moisture": moisture_results, "plate": plate_results}, f, ensure_ascii=False, indent=2)
        f.write(";\n")

    print(f"\n성공적으로 {output_file}이 업데이트 되었습니다.")


if __name__ == "__main__":
    extract_quality()
