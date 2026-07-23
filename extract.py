import os
import json
import datetime
import openpyxl

def to_num(val):
    if val is None:
        return 0
    try:
        # float/int가 섞일 수 있음
        n = float(val)
        if n.is_integer():
            return int(n)
        return n
    except (ValueError, TypeError):
        return 0

def to_str(val):
    if val is None:
        return ""
    return str(val).strip()

def format_date(val):
    if val is None:
        return ""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    # 시리얼 번호인 경우 (Excel epoch 1899-12-30)
    if isinstance(val, (int, float)):
        try:
            date_val = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=val)
            return date_val.strftime("%Y-%m-%d")
        except Exception:
            pass
    return str(val).strip()

def collect_text(ws, col, start_row, end_row):
    lines = []
    for r in range(start_row, end_row + 1):
        v = ws[f"{col}{r}"].value
        if v is not None and str(v).strip() != "":
            lines.append(str(v).rstrip())
    return lines

def collect_personnel(ws):
    rows = []
    for r in range(202, 229):
        role = ws[f"H{r}"].value
        if role is None or str(role).strip() == "":
            continue
        rows.append({
            "role": str(role).strip(),
            "prev": to_num(ws[f"J{r}"].value),
            "today": to_num(ws[f"K{r}"].value),
            "cum": to_num(ws[f"L{r}"].value)
        })
    total = {
        "prev": to_num(ws["J229"].value),
        "today": to_num(ws["K229"].value),
        "cum": to_num(ws["L229"].value)
    }
    return rows, total

def collect_equipment(ws):
    rows = []
    cur_type = None
    for r in range(202, 229):
        typ = ws[f"M{r}"].value
        spec = ws[f"N{r}"].value
        if typ is not None and str(typ).strip() != "":
            cur_type = str(typ).strip()
        if spec is None and typ is None:
            continue
        if cur_type is None:
            continue
        rows.append({
            "type": cur_type,
            "spec": str(spec).strip() if spec is not None else "",
            "prev": to_num(ws[f"O{r}"].value),
            "today": to_num(ws[f"P{r}"].value),
            "cum": to_num(ws[f"Q{r}"].value)
        })
    total = {
        "prev": to_num(ws["O229"].value),
        "today": to_num(ws["P229"].value),
        "cum": to_num(ws["Q229"].value)
    }
    
    agg_map = {}
    order = []
    for row in rows:
        t = row["type"]
        if t not in agg_map:
            agg_map[t] = {"type": t, "prev": 0, "today": 0, "cum": 0}
            order.append(t)
        agg_map[t]["prev"] += row["prev"]
        agg_map[t]["today"] += row["today"]
        agg_map[t]["cum"] += row["cum"]
    agg = [agg_map[t] for t in order]
    
    return rows, agg, total

def collect_earth(ws):
    earth_rows = {
        69: {"label": "토사", "start": "2025-02-01"},
        70: {"label": "리핑암", "start": "2025-03-10"},
        71: {"label": "발파암깎기", "start": "2025-04-05"},
        82: {"label": "흙쌓기(성토)", "start": "2025-02-01"}
    }
    earth = {}
    for r, info in earth_rows.items():
        label = info["label"]
        start = info["start"]
        earth[label] = {
            "design": to_num(ws[f"E{r}"].value),
            "prev": to_num(ws[f"F{r}"].value),
            "today": to_num(ws[f"G{r}"].value),
            "cum": to_num(ws[f"H{r}"].value),
            "start": start
        }
    return earth

def extract_excel():
    excel_dir = "작업일보"
    excel_file = None
    if not os.path.exists(excel_dir):
        print(f"에러: '{excel_dir}' 폴더가 존재하지 않습니다.")
        return
        
    for f in os.listdir(excel_dir):
        if f.endswith(".xlsx") and not f.startswith("~$"):
            excel_file = os.path.join(excel_dir, f)
            break
            
    if not excel_file:
        print("에러: '작업일보' 폴더에서 엑셀(.xlsx) 파일을 찾을 수 없습니다.")
        return
        
    print(f"엑셀 파일 읽는 중: {excel_file}")
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    results = {}
    for i in range(1, 32):
        sheet_name = f"{i:02d}"
        if sheet_name not in wb.sheetnames:
            continue
            
        ws = wb[sheet_name]
        
        plan = to_num(ws["N4"].value)
        actual = to_num(ws["O4"].value)
        diff = to_num(ws["P4"].value)
        
        p_rows, p_total = collect_personnel(ws)
        e_rows, e_agg, e_total = collect_equipment(ws)
        work_today = collect_text(ws, "B", 7, 49)
        work_tomorrow = collect_text(ws, "J", 7, 49)
        earth = collect_earth(ws)
        
        # 데이터가 입력된 유효 시트만 파싱
        if p_total["today"] > 0 or e_total["today"] > 0 or len(work_today) > 0:
            results[sheet_name] = {
                "date": format_date(ws["M5"].value),
                "weather": to_str(ws["Q5"].value),
                "progress": {
                    "plan": round(plan * 100, 2),
                    "actual": round(actual * 100, 2),
                    "diff": round(diff * 100, 2)
                },
                "earth": earth,
                "work_today": work_today,
                "work_tomorrow": work_tomorrow,
                "personnel": p_rows,
                "personnel_total": p_total,
                "equipment_detail": e_rows,
                "equipment": e_agg,
                "equipment_total": e_total
            }
            print(f"시트 {sheet_name} 파싱 완료: {results[sheet_name]['date']}")
            
    output_file = "dailyData.js"
    with open(output_file, "w", encoding="utf-8") as f:
        f.write("/**\n")
        f.write(" * 작업일보 엑셀 자동 추출 결과 데이터\n")
        f.write(" * 생성 스크립트: extract.py\n")
        f.write(" */\n")
        f.write("let DAILY_REPORTS = ")
        json_data = json.dumps(results, ensure_ascii=False, indent=2)
        f.write(json_data)
        f.write(";\n")
        
    print(f"\n성공적으로 {output_file}이 업데이트 되었습니다.")

if __name__ == "__main__":
    extract_excel()
