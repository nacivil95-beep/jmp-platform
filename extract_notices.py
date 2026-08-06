#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extract_notices.py
------------------------------------------------------
데이터관리\\알림창 폴더의 "공지사항.xlsx" 를 읽어서
저장소 루트의 notices.js 로 변환합니다.

extract.py / extract_quality.py / extract_safety.py 와 동일한 패턴:
  - 소스 엑셀 경로는 아래 SOURCE_XLSX 에서 직접 수정
  - 엑셀 파일이 없으면 에러 없이 조용히 건너뜀 (notices.js 유지)
  - auto_update.ps1 이 매일 이 스크립트를 실행

이 파일 자체를 수정할 필요는 거의 없고, 공지 내용은 항상
공지사항.xlsx 파일에서만 수정하면 됩니다.
"""

import os
import sys
import io
from datetime import datetime, date

import openpyxl
from PIL import Image as PILImage

# ── 환경 설정 ────────────────────────────────────────────────────────────
# 이 스크립트가 있는 폴더(=git 저장소 루트) 기준 출력 경로
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# 공지사항 엑셀 원본 경로: 저장소 안의 데이터관리\알림창\공지사항.xlsx
# (저장소 자체를 다른 폴더/PC로 옮겨도 이 경로는 자동으로 따라가므로 수정할 필요 없음)
SOURCE_XLSX = os.path.join(SCRIPT_DIR, "데이터관리", "알림창", "공지사항.xlsx")

OUTPUT_JS = os.path.join(SCRIPT_DIR, "notices.js")
ASSET_DIR = os.path.join(SCRIPT_DIR, "assets", "notice")  # 셀에 직접 삽입한 그림 저장 위치

# ── notices.js 필드 정의 ────────────────────────────────────────────────
ALL_FIELDS = ["id", "team", "level", "title", "body", "image", "startDate", "endDate", "pinned"]
REQUIRED_FOR_HEADER_DETECTION = {"id", "team", "title", "body"}

JS_HEADER = """/**
 * 사이트 접속 시 뜨는 공지사항 팝업 데이터
 * ------------------------------------------------------------------------
 * 이 파일은 공지사항.xlsx 에서 extract_notices.py 스크립트로 자동 생성됩니다.
 * 직접 이 파일을 수정하지 말고, 엑셀에서 수정한 뒤 auto_update 를 실행하세요.
 *
 *   id         : 고유 값
 *   team       : 작성 팀/구분 이름 - 카드에 태그로 표시됨
 *   level      : "info"(파랑) | "warning"(주황) | "danger"(빨강)
 *   title      : 공지 제목
 *   body       : 공지 내용 (줄바꿈 \\n, 굵게 <b>...</b>)
 *   image      : (선택) 이미지 경로
 *   startDate  : (선택) "YYYY-MM-DD"
 *   endDate    : (선택) "YYYY-MM-DD"
 *   pinned     : true 면 항상 상단 고정
 *
 * 원본: {source}
 * 생성 시각: {timestamp}
 */
"""


def cell_to_str(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def find_header_map(ws):
    for row in ws.iter_rows(min_row=1, max_row=30):
        raw_values = [cell_to_str(c.value) for c in row]
        norm_values = [v.strip().lower() for v in raw_values]

        found = {}
        for idx, norm in enumerate(norm_values):
            for field in ALL_FIELDS:
                if norm == field.lower():
                    found[field] = idx + 1

        if REQUIRED_FOR_HEADER_DETECTION.issubset(found.keys()):
            return row[0].row, found

    raise ValueError(
        "헤더 행을 찾을 수 없습니다. 시트 안에 최소한 "
        "'id, team, title, body' 열 이름이 정확히 있는지 확인해주세요."
    )


def extract_embedded_images(ws, image_col_idx, asset_dir, id_by_row):
    result = {}
    if image_col_idx is None:
        return result

    images = getattr(ws, "_images", [])
    if not images:
        return result

    os.makedirs(asset_dir, exist_ok=True)

    for img in images:
        anchor = img.anchor
        from_row0 = anchor._from.row
        from_col0 = anchor._from.col
        excel_row = from_row0 + 1
        excel_col = from_col0 + 1

        if excel_col != image_col_idx:
            continue
        if excel_row not in id_by_row:
            continue

        data = img.ref
        if hasattr(data, "read"):
            data.seek(0)
            raw = data.read()
        else:
            raw = data

        pil_img = PILImage.open(io.BytesIO(raw))
        ext = (pil_img.format or "PNG").lower()
        if ext == "jpeg":
            ext = "jpg"

        notice_id = id_by_row[excel_row] or f"row{excel_row}"
        filename = f"{notice_id}.{ext}"
        out_path = os.path.join(asset_dir, filename)
        with open(out_path, "wb") as f:
            f.write(raw)

        # notices.js 에는 사이트 기준 상대경로(assets/notice/...)로 기록
        rel_path = os.path.relpath(out_path, SCRIPT_DIR).replace("\\", "/")
        result[excel_row] = rel_path

    return result


def load_notices(xlsx_path, sheet_name=None, asset_dir=ASSET_DIR):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    header_row, col_map = find_header_map(ws)
    data_start = header_row + 1
    image_col_idx = col_map.get("image")

    def get(row_cells, field):
        col = col_map.get(field)
        if not col:
            return ""
        return cell_to_str(row_cells[col - 1].value)

    id_by_row = {}
    id_col = col_map.get("id")
    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row):
        rid = cell_to_str(row[id_col - 1].value) if id_col else ""
        if rid:
            id_by_row[row[0].row] = rid

    embedded_images = extract_embedded_images(ws, image_col_idx, asset_dir, id_by_row)

    notices = []
    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row):
        rid = get(row, "id")
        if not rid:
            continue

        title = get(row, "title")
        body_raw = get(row, "body")
        if not title and not body_raw:
            # title/body가 둘 다 비어있으면 설명/주석 줄일 가능성이 높아 건너뜀
            continue

        excel_row = row[0].row
        image_val = embedded_images.get(excel_row) or get(row, "image")
        body = body_raw.replace("\r\n", "\n").replace("\r", "\n")

        pinned_raw = get(row, "pinned").strip().lower()
        pinned = pinned_raw in ("true", "1", "yes", "y", "o")

        notice = {
            "id": rid,
            "team": get(row, "team"),
            "level": get(row, "level") or "info",
            "title": title,
            "body": body,
        }
        if image_val:
            notice["image"] = image_val
        notice["startDate"] = get(row, "startDate")
        notice["endDate"] = get(row, "endDate")
        notice["pinned"] = pinned

        notices.append(notice)

    return notices


def js_string_literal(s):
    escaped = (
        s.replace("\\", "\\\\")
         .replace('"', '\\"')
         .replace("\n", "\\n")
    )
    return f'"{escaped}"'


def notice_to_js(notice, indent="   "):
    order = ["id", "team", "level", "title", "image", "body", "startDate", "endDate", "pinned"]
    lines = [f"{indent}{{"]
    field_indent = indent + " "
    for key in order:
        if key not in notice:
            continue
        val = notice[key]
        if key == "pinned":
            js_val = "true" if val else "false"
        else:
            js_val = js_string_literal(val)
        lines.append(f"{field_indent}{key}: {js_val},")
    lines[-1] = lines[-1].rstrip(",")
    lines.append(f"{indent}}}")
    return "\n".join(lines)


def build_js(notices, source_path):
    header = JS_HEADER.format(
        source=source_path,
        timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )
    items = ",\n".join(notice_to_js(n) for n in notices)
    body = f"let SITE_NOTICES = [\n{items}\n];\n" if notices else "let SITE_NOTICES = [\n];\n"
    return header + body


def main():
    if not os.path.exists(SOURCE_XLSX):
        # extract_quality.py 와 동일하게: 파일이 없으면 조용히 건너뜀 (에러 아님)
        print(f"공지사항.xlsx 를 찾을 수 없어 건너뜁니다: {SOURCE_XLSX}")
        sys.exit(0)

    try:
        notices = load_notices(SOURCE_XLSX)
    except Exception as e:
        # 형식이 깨져도 auto_update 전체가 죽지 않도록 notices.js는 건드리지 않고 종료
        print(f"공지사항.xlsx 처리 중 오류 발생, notices.js는 갱신하지 않습니다: {e}")
        sys.exit(0)

    js_text = build_js(notices, SOURCE_XLSX)

    with open(OUTPUT_JS, "w", encoding="utf-8") as f:
        f.write(js_text)

    print(f"완료: {len(notices)}개 공지 -> {OUTPUT_JS}")


if __name__ == "__main__":
    main()
